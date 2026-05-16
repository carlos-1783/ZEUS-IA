"""Importación CSV/XLSX de clientes al CRM de oficina."""

from __future__ import annotations

import csv
import io
import logging
import re
import unicodedata
from typing import Any, Dict, List, Optional, Set, Tuple

from email_validator import EmailNotValidError, validate_email
from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.customer import Customer
from app.models.user import User
from app.schemas.crm_import import CrmImportColumnMapping
import services.crm_office_service as crm_svc

logger = logging.getLogger(__name__)

MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_ROWS = 5000
PREVIEW_ROWS = 15

FIELD_ALIASES: Dict[str, List[str]] = {
    "name": [
        "nombre",
        "cliente",
        "full_name",
        "fullname",
        "name",
        "razon_social",
        "razón_social",
        "empresa",
        "contacto",
        "denominacion",
        "denominación",
    ],
    "email": ["email", "correo", "e_mail", "mail", "correo_electronico"],
    "phone": [
        "telefono",
        "teléfono",
        "tel",
        "movil",
        "móvil",
        "mobile",
        "phone",
        "celular",
        "telefono_movil",
    ],
    "notes": ["notas", "observaciones", "notes", "comentarios", "observacion", "descripcion"],
    "tax_id": ["nif", "cif", "dni", "tax_id", "nif_cif", "identificacion", "id_fiscal"],
}


def _norm_header(value: str) -> str:
    if not value:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower().replace(" ", "_").replace("-", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def _norm_phone(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    digits = re.sub(r"\D", "", str(value))
    if not digits:
        return None
    if len(digits) > 15:
        digits = digits[-15:]
    if len(digits) >= 9 and not digits.startswith("+"):
        return f"+{digits}"
    return digits if digits.startswith("+") else f"+{digits}"


def _validate_email(value: Optional[str]) -> Optional[str]:
    if not value or not str(value).strip():
        return None
    raw = str(value).strip().lower()
    try:
        result = validate_email(raw, check_deliverability=False)
        return result.normalized
    except EmailNotValidError:
        return None


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def detect_column_mapping(columns: List[str]) -> CrmImportColumnMapping:
    normalized = {_norm_header(c): c for c in columns if c}
    mapping: Dict[str, Optional[str]] = {
        "name": None,
        "email": None,
        "phone": None,
        "notes": None,
        "tax_id": None,
    }
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            key = _norm_header(alias)
            if key in normalized:
                mapping[field] = normalized[key]
                break
    return CrmImportColumnMapping(**mapping)


def parse_upload(content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    if len(content) > MAX_FILE_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo supera el límite de {MAX_FILE_BYTES // (1024 * 1024)} MB.",
        )
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return _parse_csv(content)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Formato no soportado. Usa CSV o XLSX.",
    )


def _parse_csv(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se pudo leer el CSV (codificación no reconocida).",
        )

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El CSV no tiene cabeceras.")

    columns = [c.strip() for c in reader.fieldnames if c and str(c).strip()]
    rows: List[Dict[str, Any]] = []
    for raw in reader:
        if len(rows) >= MAX_ROWS:
            break
        row = {_cell_str(k): _cell_str(raw.get(k, "")) for k in columns}
        rows.append(row)
    return columns, rows


def _parse_xlsx(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("xlsx parse failed: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se pudo leer el archivo Excel. Comprueba que sea un .xlsx válido.",
        ) from exc

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel no tiene hojas.")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel está vacío.")

    columns: List[str] = []
    for idx, cell in enumerate(header_row):
        label = _cell_str(cell) or f"column_{idx + 1}"
        columns.append(label)

    rows: List[Dict[str, Any]] = []
    for values in rows_iter:
        if len(rows) >= MAX_ROWS:
            break
        row: Dict[str, Any] = {}
        empty = True
        for col_idx, col_name in enumerate(columns):
            val = values[col_idx] if col_idx < len(values) else None
            s = _cell_str(val)
            if s:
                empty = False
            row[col_name] = s
        if not empty:
            rows.append(row)
    wb.close()
    return columns, rows


def _row_is_empty(mapped: Dict[str, Optional[str]]) -> bool:
    return not any(v and str(v).strip() for v in mapped.values())


def _apply_mapping(row: Dict[str, Any], mapping: CrmImportColumnMapping) -> Dict[str, Optional[str]]:
    def pick(field: Optional[str]) -> Optional[str]:
        if not field or field not in row:
            return None
        v = _cell_str(row.get(field, ""))
        return v or None

    name = pick(mapping.name)
    email_raw = pick(mapping.email)
    phone_raw = pick(mapping.phone)
    notes = pick(mapping.notes)
    tax_id = pick(mapping.tax_id)

    email = _validate_email(email_raw) if email_raw else None
    phone = _norm_phone(phone_raw)

    if tax_id:
        tax_id = re.sub(r"\s+", "", tax_id)[:50] or None

    return {
        "name": name[:100] if name else None,
        "email": email,
        "phone": phone[:20] if phone else None,
        "notes": notes,
        "tax_id": tax_id,
    }


def build_preview(
    content: bytes,
    filename: str,
) -> Tuple[List[str], CrmImportColumnMapping, List[Dict[str, Any]], int]:
    columns, rows = parse_upload(content, filename)
    if not columns:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se detectaron columnas.")
    suggested = detect_column_mapping(columns)
    if not suggested.name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se detectó columna de nombre. Añade una cabecera como «nombre» o «cliente».",
        )

    preview: List[Dict[str, Any]] = []
    for row in rows[:PREVIEW_ROWS]:
        mapped = _apply_mapping(row, suggested)
        preview.append({**mapped, "_raw": row})

    return columns, suggested, preview, len(rows)


def _existing_email_phone_sets(db: Session, company_id: Optional[int]) -> Tuple[Set[str], Set[str]]:
    emails: Set[str] = set()
    phones: Set[str] = set()
    q = db.query(Customer.email, Customer.phone)
    if company_id is not None:
        q = q.filter(Customer.company_id == company_id)
    else:
        q = q.filter(Customer.company_id.is_(None))
    for email, phone in q.all():
        if email:
            emails.add(str(email).strip().lower())
        norm = _norm_phone(phone)
        if norm:
            phones.add(norm)
    return emails, phones


def import_customers(
    db: Session,
    user: User,
    *,
    content: bytes,
    filename: str,
    mapping: CrmImportColumnMapping,
) -> Dict[str, Any]:
    if not mapping.name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Debes indicar la columna de nombre en el mapeo.",
        )

    columns, rows = parse_upload(content, filename)
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo no tiene filas de datos.")

    company_id = crm_svc.primary_company_id(db, user)
    if company_id is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Se requiere una empresa asociada para importar clientes.",
        )

    existing_emails, existing_phones = _existing_email_phone_sets(db, company_id)
    batch_emails: Set[str] = set()
    batch_phones: Set[str] = set()

    imported = 0
    skipped_empty = 0
    skipped_duplicates = 0
    skipped_invalid = 0
    errors: List[str] = []

    try:
        for idx, row in enumerate(rows, start=2):
            mapped = _apply_mapping(row, mapping)
            if _row_is_empty(mapped):
                skipped_empty += 1
                continue

            name = mapped.get("name")
            if not name or len(name) < 2:
                skipped_invalid += 1
                if len(errors) < 20:
                    errors.append(f"Fila {idx}: nombre obligatorio (mín. 2 caracteres).")
                continue

            email = mapped.get("email")
            phone = mapped.get("phone")

            if email and email in existing_emails:
                skipped_duplicates += 1
                continue
            if email and email in batch_emails:
                skipped_duplicates += 1
                continue
            if phone and phone in existing_phones:
                skipped_duplicates += 1
                continue
            if phone and phone in batch_phones:
                skipped_duplicates += 1
                continue

            customer = Customer(
                name=name,
                email=email,
                phone=phone,
                tax_id=mapped.get("tax_id"),
                notes=mapped.get("notes"),
                is_active=True,
                is_company=True,
                company_id=company_id,
                owner_user_id=user.id,
            )
            db.add(customer)
            imported += 1
            if email:
                batch_emails.add(email)
                existing_emails.add(email)
            if phone:
                batch_phones.add(phone)
                existing_phones.add(phone)

        if imported == 0:
            db.rollback()
            return {
                "success": True,
                "imported": 0,
                "skipped": skipped_empty + skipped_duplicates + skipped_invalid,
                "skipped_empty": skipped_empty,
                "skipped_duplicates": skipped_duplicates,
                "skipped_invalid": skipped_invalid,
                "errors": errors,
                "message": "No se importó ningún cliente (todas las filas vacías o duplicadas).",
            }

        db.flush()
        crm_svc.log_activity(
            db,
            company_id=company_id,
            user_id=user.id,
            customer_id=None,
            record_id=None,
            action="customers_imported",
            summary=f"Importación masiva: {imported} cliente(s) desde {filename}",
            payload={
                "filename": filename,
                "imported": imported,
                "skipped_duplicates": skipped_duplicates,
                "skipped_empty": skipped_empty,
                "skipped_invalid": skipped_invalid,
            },
            commit=False,
        )
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        logger.exception("import_customers failed")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al importar: {exc}",
        ) from exc

    skipped = skipped_empty + skipped_duplicates + skipped_invalid
    return {
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "skipped_empty": skipped_empty,
        "skipped_duplicates": skipped_duplicates,
        "skipped_invalid": skipped_invalid,
        "errors": errors,
        "message": f"Importados {imported} cliente(s). Omitidas {skipped} fila(s).",
    }
