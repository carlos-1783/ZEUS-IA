"""Generación real de Excel modelo 303 (modelo_303_v1)."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict

logger = logging.getLogger(__name__)


def generate_model_303_xlsx_v1(
    *,
    output_path: Path,
    company_name: str,
    period: str,
    model_data: Dict[str, Any],
) -> str:
    """Escribe XLSX del modelo 303. Devuelve ruta absoluta."""
    try:
        from openpyxl import Workbook  # pyright: ignore[reportMissingModuleSource]
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl no disponible para generar modelo 303") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_path.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo 303"

    bold = Font(bold=True)
    ws["A1"] = "Modelo 303 — IVA trimestral"
    ws["A1"].font = bold
    ws["A2"] = "Empresa"
    ws["B2"] = company_name
    ws["A3"] = "Periodo"
    ws["B3"] = period

    ws["A5"] = "Concepto"
    ws["B5"] = "Importe (€)"
    ws["A5"].font = bold
    ws["B5"].font = bold

    rows = [
        ("Base imponible (facturas)", model_data.get("base_imponible", 0)),
        ("IVA devengado / repercutido", model_data.get("iva_devengado", 0)),
        ("IVA soportado (gastos)", model_data.get("iva_soportado", 0)),
        ("Resultado (a ingresar / compensar)", model_data.get("resultado", 0)),
    ]
    row_idx = 6
    for label, value in rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=round(float(value or 0), 2))
        row_idx += 1

    # Desglose por tipo de IVA si existe
    breakdown = model_data.get("vat_breakdown") or {}
    if breakdown:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="Desglose IVA devengado").font = bold
        row_idx += 1
        for rate, amount in sorted(breakdown.items(), key=lambda x: str(x[0])):
            ws.cell(row=row_idx, column=1, value=f"Tipo {rate}%")
            ws.cell(row=row_idx, column=2, value=round(float(amount or 0), 2))
            row_idx += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    wb.save(xlsx_path)
    logger.info("Model 303 XLSX generated: %s", xlsx_path)
    return str(xlsx_path.resolve())
